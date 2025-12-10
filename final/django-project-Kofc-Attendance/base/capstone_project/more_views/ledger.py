# ledger.py – FINAL PATCHED: Name search fixed + deduplicated + 3-tier access
from capstone_project.models import blockchain, Donation
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, FileResponse
from django.utils.html import escape
from django.views.decorators.cache import never_cache
from django.core.paginator import Paginator
from django.contrib import messages
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from django.urls import reverse
from openpyxl.utils import get_column_letter
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from dateutil import parser as date_parser
import io, logging, re

logger = logging.getLogger(__name__)

# ========================
# PDF Receipt (unchanged)
# ========================
def generate_receipt_pdf(donation):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.setFont("Helvetica-Bold", 16)
    p.drawString(inch, height - inch, "Knights of Columbus - Donation Receipt")

    donor_name = f"{donation.first_name or ''} {donation.middle_initial or ''} {donation.last_name or ''}".strip() or "Anonymous Donor"
    email = donation.email or "N/A"
    event_name = donation.event.name if donation.event else "General Donation"

    data = {
        'transaction_id': donation.transaction_id,
        'donor_name': donor_name,
        'email': email,
        'amount': donation.amount,
        'donation_date': donation.donation_date,
        'payment_method': donation.payment_method.capitalize(),
        'event_name': event_name,
        'status': donation.get_status_display(),
        'block_index': "Pending" if donation.status != 'completed' else "Recorded",
    }

    y = height - 2 * inch
    p.setFont("Helvetica", 12)
    for key, value in data.items():
        p.drawString(inch, y, f"{key.replace('_', ' ').title()}: {value}")
        y -= 0.25 * inch

    p.drawString(inch, inch, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    p.drawString(inch, inch - 0.25 * inch, "Thank you for your support! Verify on blockchain ledger.")

    p.save()
    buffer.seek(0)
    return buffer

def sanitize_input(value, max_length=100):
    if not value:
        return ''
    sanitized = re.sub(r'[<>";\'\\]', '', str(value))
    sanitized = escape(sanitized)
    return sanitized[:max_length]

@login_required
def download_receipt(request, donation_id):
    donation = get_object_or_404(Donation, id=donation_id)
    
    is_admin_or_officer = request.user.role in ['admin', 'officer']
    is_own_donation = request.user.is_authenticated and request.user.email == donation.email
    
    if not (is_admin_or_officer or is_own_donation):
        messages.info(request, "You can request the receipt via email if this is your donation.")
        return redirect('request_receipt', donation_id=donation_id)
    
    if request.user.role == 'officer' and donation.council and donation.council != request.user.council:
        messages.error(request, "You can only download receipts for your council's donations.")
        return redirect('blockchain')
    
    logger.info(f"User {request.user.username} (role: {request.user.role}) downloaded receipt for donation {donation_id}")
    
    pdf_buffer = generate_receipt_pdf(donation)
    return FileResponse(
        pdf_buffer,
        as_attachment=True,
        filename=f"receipt_{donation.transaction_id}.pdf"
    )

# ========================
# Privacy Helpers
# ========================
def mask_name(name):
    if not name:
        return 'N/A'
    parts = name.split()
    masked_parts = []
    for part in parts:
        if len(part) <= 2:
            masked_parts.append(part)
        else:
            middle_len = len(part) - 2
            masked_parts.append(part[0] + '*' * middle_len + part[-1])
    return ' '.join(masked_parts)

def mask_email(email):
    if not email or '@' not in email:
        return 'N/A'
    local, domain = email.split('@', 1)
    if len(local) <= 2:
        masked_local = local
    else:
        middle_len = len(local) - 2
        masked_local = local[0] + '*' * middle_len + local[-1]
    return masked_local + '@' + domain

# ========================
# Data Normalizer (with raw_donor for search)
# ========================
def normalize_blockchain_data(full_chain, pending_transactions):
    def normalize_tx(tx):
        # Extract raw values
        tx_id = getattr(tx, 'id', None) if not isinstance(tx, dict) else tx.get('id')
        transaction_id = getattr(tx, 'transaction_id', '') if not isinstance(tx, dict) else tx.get('transaction_id', '')
        donor_raw = (tx.get_display_name() if hasattr(tx, 'get_display_name') 
                     else getattr(tx, 'donor', '') if not isinstance(tx, dict) 
                     else tx.get('donor_name', tx.get('donor', '')))
        email = getattr(tx, 'email', 'N/A') if not isinstance(tx, dict) else tx.get('email', 'N/A')
        amount = getattr(tx, 'amount', 0.0) if not isinstance(tx, dict) else tx.get('amount', 0.0)
        donation_date = getattr(tx, 'donation_date', None) if not isinstance(tx, dict) else tx.get('donation_date', None)
        payment_method = getattr(tx, 'payment_method', 'N/A') if not isinstance(tx, dict) else tx.get('payment_method', 'N/A')
        is_anonymous = getattr(tx, 'is_anonymous', False) if not isinstance(tx, dict) else tx.get('is_anonymous', donor_raw == 'Anonymous Donor')
        submitted_by = getattr(tx, 'submitted_by', 'N/A') if not isinstance(tx, dict) else tx.get('submitted_by', 'N/A')
        reviewed_by = getattr(tx, 'reviewed_by', 'N/A') if not isinstance(tx, dict) else tx.get('reviewed_by', 'N/A')

        # Fix ID if missing
        if tx_id is None and transaction_id:
            try:
                donation = Donation.objects.get(transaction_id=transaction_id)
                tx_id = donation.id
            except (Donation.DoesNotExist, Donation.MultipleObjectsReturned):
                tx_id = None

        # Parse amount
        if isinstance(amount, str):
            try:
                amount = float(amount.replace('₱', '').replace(',', '').strip())
            except ValueError:
                amount = 0.0
        elif amount is None:
            amount = 0.0

        # Parse date
        if donation_date and isinstance(donation_date, str):
            try:
                donation_date = datetime.strptime(donation_date, '%Y-%m-%d').date()
            except ValueError:
                donation_date = None

        # Build tx dict
        tx_dict = {
            'id': tx_id,
            'transaction_id': transaction_id,
            'amount': amount,
            'donation_date': donation_date,
            'payment_method': payment_method,
            'submitted_by': submitted_by,
            'reviewed_by': reviewed_by,
        }

        # Donor & Email (with raw for search)
        if is_anonymous:
            tx_dict['donor'] = "Anonymous Donor"
            tx_dict['email'] = "N/A"
            tx_dict['raw_donor'] = ""
        else:
            tx_dict['donor'] = mask_name(donor_raw)
            tx_dict['email'] = mask_email(email)
            tx_dict['raw_donor'] = donor_raw.lower()

        return tx_dict

    # Normalize chain
    for block in full_chain:
        if not isinstance(block, dict):
            block = {
                'index': block.index,
                'timestamp': block.timestamp,
                'transactions': block.transactions if hasattr(block, 'transactions') else [],
                'proof': block.proof,
                'previous_hash': block.previous_hash,
                'hash': block.hash
            }
        if isinstance(block.get('timestamp'), str):
            try:
                block['timestamp'] = date_parser.parse(block['timestamp'])
            except ValueError:
                block['timestamp'] = None
        block['transactions'] = [normalize_tx(tx) for tx in block.get('transactions', [])]

    # Normalize pending
    pending_transactions = [normalize_tx(tx) for tx in pending_transactions]

    return full_chain, pending_transactions

# ========================
# Shared Transaction List Builder
# ========================
def build_transaction_list(full_chain, pending_transactions, filters, sort_by, user):
    search_name = filters.get('search_name', '').lower()
    search_tx = filters.get('search_tx', '').lower()
    date_from = filters.get('date_from')
    date_to = filters.get('date_to')
    amount_min = filters.get('amount_min', '')
    amount_max = filters.get('amount_max', '')
    method = filters.get('method', '')

    def matches(tx):
        # Search by Transaction ID
        if search_tx and search_tx not in str(tx.get('transaction_id', '')).lower():
            return False

        # Search by Donor Name (tiered)
        if search_name:
            if user.role in ['admin', 'officer']:
                if search_name not in tx.get('raw_donor', ''):
                    return False
            else:
                if search_name not in str(tx.get('donor', '')).lower():
                    return False

        # Date range
        donation_date = tx.get('donation_date')
        if date_from and donation_date and donation_date < datetime.strptime(date_from, '%Y-%m-%d').date():
            return False
        if date_to and donation_date and donation_date > datetime.strptime(date_to, '%Y-%m-%d').date():
            return False

        # Amount
        try:
            amount = float(tx.get('amount', 0))
            if amount_min and amount < float(amount_min):
                return False
            if amount_max and amount > float(amount_max):
                return False
        except (ValueError, TypeError):
            pass

        # Method
        if method and method != tx.get('payment_method', ''):
            return False

        return True

    # Filter
    filtered_chain = [
        {**block, 'transactions': [tx for tx in block.get('transactions', []) if matches(tx)]}
        for block in full_chain
        if any(matches(tx) for tx in block.get('transactions', []))
    ]
    filtered_pending = [tx for tx in pending_transactions if matches(tx)]

    # Flatten
    all_tx = []
    max_idx = max((b.get('index', 0) for b in full_chain), default=0)

    for block in filtered_chain:
        for tx in block['transactions']:
            t = tx.copy()
            t.update({
                'block_index': block['index'],
                'block_timestamp': block['timestamp'],
                'previous_hash': block['previous_hash'],
                'block_hash': block['hash'],
                'proof': block['proof'],
                '_sort_key': block['index']
            })
            all_tx.append(t)

    for tx in filtered_pending:
        t = tx.copy()
        t.update({
            'block_index': None,
            'block_timestamp': None,
            'previous_hash': 'N/A',
            'block_hash': 'N/A',
            'proof': 'N/A',
            '_sort_key': max_idx + 1
        })
        all_tx.append(t)

    # Sort
    if sort_by == 'recent_to_oldest':
        all_tx.sort(key=lambda x: x['_sort_key'], reverse=True)
    elif sort_by == 'oldest_to_recent':
        all_tx.sort(key=lambda x: x['_sort_key'])
    elif sort_by == 'highest_amount':
        all_tx.sort(key=lambda x: x.get('amount', 0), reverse=True)
    elif sort_by == 'lowest_amount':
        all_tx.sort(key=lambda x: x.get('amount', 0))

    return all_tx

# ========================
# Main View
# ========================
@never_cache
def get_blockchain_data(request):
    try:
        full_chain = blockchain.get_chain()
        if not blockchain.is_chain_valid():
            messages.error(request, "Blockchain data is corrupted. Contact support.")
            return redirect('donations')

        pending_transactions = blockchain.pending_transactions
        full_chain, pending_transactions = normalize_blockchain_data(full_chain, pending_transactions)

        # === Filters (public + auth) ===
        search_name = sanitize_input(request.GET.get('search_name', ''), 50).lower() if request.user.is_authenticated else ''
        search_tx = request.GET.get('search_tx', '').strip().lower() if request.user.is_authenticated else ''
        date_from_str = request.GET.get('date_from', '').strip()
        date_to_str = request.GET.get('date_to', '').strip()
        amount_min = request.GET.get('amount_min', '').strip()
        amount_max = request.GET.get('amount_max', '').strip()
        method = request.GET.get('method', '').strip()
        sort = request.GET.get('sort', 'recent_to_oldest')

        # === Date Validation ===
        date_from = date_to = None
        date_error = False

        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid 'Date From' format.")
                date_error = True

        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid 'Date To' format.")
                date_error = True

        if not date_error and date_from and date_to and date_from > date_to:
            messages.error(request, "Error: 'Date From' cannot be later than 'Date To'.")
            date_error = True

        if date_error:
            query_params = request.GET.copy()
            for key in ['date_from', 'date_to']:
                query_params.pop(key, None)
            return redirect(f"{reverse('blockchain')}?{query_params.urlencode()}")

        # === Build List ===
        filters = {
            'search_name': search_name,
            'search_tx': search_tx,
            'date_from': date_from_str,
            'date_to': date_to_str,
            'amount_min': amount_min,
            'amount_max': amount_max,
            'method': method,
        }
        all_transactions = build_transaction_list(full_chain, pending_transactions, filters, sort, request.user)

        # === Pagination ===
        paginator = Paginator(all_transactions, 10)
        page_obj = paginator.get_page(request.GET.get('page', 1))

        return render(request, 'blockchain.html', {
            'page_obj': page_obj,
            'total_blocks': len([b for b in full_chain if b.get('transactions')]),
            'total_transactions': len(all_transactions),
            'pending_transactions': pending_transactions,
            'admin_officer_roles': ['admin', 'officer'],
        })

    except Exception as e:
        logger.error(f"Error fetching blockchain data: {str(e)}")
        messages.error(request, "Unable to retrieve blockchain data.")
        return redirect('donations')

# ========================
# Download Ledger
# ========================
def download_ledger(request):
    if not request.user.is_authenticated:
        messages.error(request, "Please log in to download the ledger.")
        return redirect('login')

    if request.user.role not in ['admin', 'officer']:
        messages.error(request, "You do not have permission to download the full ledger.")
        return redirect(reverse('blockchain'))

    try:
        full_chain = blockchain.get_chain()
        if not blockchain.is_chain_valid():
            messages.error(request, "Blockchain integrity check failed. Cannot generate ledger.")
            return redirect(reverse('blockchain'))

        pending_transactions = blockchain.pending_transactions
        full_chain, pending_transactions = normalize_blockchain_data(full_chain, pending_transactions)

        filters = {
            'search_name': request.GET.get('search_name', '').lower(),
            'search_tx': request.GET.get('search_tx', '').lower(),
            'date_from': request.GET.get('date_from'),
            'date_to': request.GET.get('date_to'),
            'amount_min': request.GET.get('amount_min', ''),
            'amount_max': request.GET.get('amount_max', ''),
            'method': request.GET.get('method', ''),
        }
        sort = request.GET.get('sort', 'recent_to_oldest')

        all_transactions = build_transaction_list(full_chain, pending_transactions, filters, sort, request.user)

        # === Excel Generation ===
        wb = Workbook()
        ws = wb.active
        ws.title = "Donation Ledger"

        headers = ['Block Index', 'Block Timestamp', 'Previous Hash', 'Block Hash', 'Proof',
                   'Amount', 'Transaction ID', 'Donor', 'Email', 'Date', 'Method',
                   'Submitted By', 'Reviewed By']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for tx in all_transactions:
            donation_date = tx.get('donation_date')
            if isinstance(donation_date, str):
                try:
                    donation_date = datetime.strptime(donation_date, '%Y-%m-%d').date()
                except:
                    donation_date = None

            amount = float(tx.get('amount', 0)) if tx.get('amount') not in (None, '') else 0.0
            block_timestamp = tx.get('block_timestamp')
            if block_timestamp and block_timestamp.tzinfo:
                block_timestamp = block_timestamp.replace(tzinfo=None)

            ws.append([
                tx.get('block_index', 'Pending') if tx.get('block_index') is not None else 'Pending',
                block_timestamp if block_timestamp else '',
                tx.get('previous_hash', 'N/A'),
                tx.get('block_hash', 'N/A'),
                tx.get('proof', 'N/A'),
                amount,
                tx.get('transaction_id', ''),
                tx.get('donor', 'N/A'),
                tx.get('email', 'N/A'),
                donation_date,
                tx.get('payment_method', 'N/A'),
                tx.get('submitted_by', 'N/A'),
                tx.get('reviewed_by', 'N/A')
            ])

        # Formatting
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = wrap_alignment

        for cell in ws['F'][1:]:
            cell.number_format = '"₱"#,##0.00'
        for cell in ws['J'][1:]:
            if cell.value:
                cell.number_format = 'YYYY-MM-DD'
        for cell in ws['B'][1:]:
            if cell.value:
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'

        for column_cells in ws.columns:
            max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
            adjusted_width = min(max_length + 4, 60)
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="donation_ledger_{timestamp}.xlsx"'
        return response

    except Exception as e:
        logger.error(f"Ledger download failed: {str(e)}")
        messages.error(request, "Failed to generate ledger. Please try again later.")
        return redirect(reverse('blockchain'))